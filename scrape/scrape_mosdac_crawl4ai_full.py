import os
import json
import asyncio
import sys
import subprocess
import fitz  # PyMuPDF
import openpyxl
from bs4 import BeautifulSoup
from crawl4ai import AsyncWebCrawler, BFSDeepCrawlStrategy

def install_playwright_browsers():
    """Checks and installs Playwright browsers if they are missing."""
    print("🔧 Checking Playwright browser installation...")
    try:
        # This command checks if browsers are installed and installs them if not.
        subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            check=True,
            capture_output=True,
            text=True
        )
        print("✅ Playwright browsers are installed.")
    except (subprocess.CalledProcessError, FileNotFoundError):
        print("❌ Error during Playwright setup. Please run this command in your terminal:")
        print(f"   playwright install chromium")
        sys.exit(1) # Exit if the core dependency is missing

async def run_the_crawl():
    """Configures and runs the asynchronous crawler, ensuring files are saved."""
    output_dir = "mosdac_crawl_all"
    crawl_depth = 2  # Reduced depth to avoid too many requests
    
    # Create output directory structure
    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(os.path.join(output_dir, "html"), exist_ok=True)
    os.makedirs(os.path.join(output_dir, "files"), exist_ok=True)

    # Initialize crawler with proper configuration
    crawler = AsyncWebCrawler(
        headless=True,
        verbose=True
    )
    
    try:
        await crawler.start()
        
        # Start with the main page
        print(f"🚀 Starting crawl of 'https://www.mosdac.gov.in'. This may take several minutes...")
        
        # First, crawl the main page
        result = await crawler.arun(
            url="https://www.mosdac.gov.in",
            bypass_cache=True,
            process_iframes=True,
            remove_overlay_elements=True,
            simulate_user=True,
            override_navigator=True
        )
        
        if result.success:
            print(f"✅ Successfully crawled main page")
            
            # Save the main page HTML
            main_html_path = os.path.join(output_dir, "html", "main_page.html")
            with open(main_html_path, "w", encoding="utf-8") as f:
                f.write(result.html)
            
            # Save the cleaned content
            main_content_path = os.path.join(output_dir, "main_page_content.txt")
            with open(main_content_path, "w", encoding="utf-8") as f:
                f.write(result.cleaned_html or result.html)
            
            # Extract links for deeper crawling
            soup = BeautifulSoup(result.html, 'html.parser')
            links = []
            base_url = "https://www.mosdac.gov.in"
            
            for link in soup.find_all('a', href=True):
                href = link['href']
                if href.startswith('/'):
                    full_url = base_url + href
                elif href.startswith('http') and 'mosdac.gov.in' in href:
                    full_url = href
                else:
                    continue
                
                if full_url not in links and full_url != "https://www.mosdac.gov.in":
                    links.append(full_url)
            
            print(f"📋 Found {len(links)} internal links to crawl")
            
            # Crawl additional pages (limit to avoid overwhelming the server)
            crawled_count = 0
            max_additional_pages = 20  # Limit to avoid too many requests
            
            for i, link in enumerate(links[:max_additional_pages]):
                try:
                    print(f"🔍 Crawling page {i+1}/{min(len(links), max_additional_pages)}: {link}")
                    
                    page_result = await crawler.arun(
                        url=link,
                        bypass_cache=True,
                        process_iframes=True,
                        remove_overlay_elements=True,
                        simulate_user=True,
                        override_navigator=True,
                        delay_before_return_html=2.0  # Wait a bit for page to load
                    )
                    
                    if page_result.success:
                        # Save HTML file
                        filename = f"page_{i+1}.html"
                        html_path = os.path.join(output_dir, "html", filename)
                        with open(html_path, "w", encoding="utf-8") as f:
                            f.write(page_result.html)
                        
                        # Check for downloadable files (PDFs, Excel, etc.)
                        page_soup = BeautifulSoup(page_result.html, 'html.parser')
                        for file_link in page_soup.find_all('a', href=True):
                            file_href = file_link['href']
                            if any(ext in file_href.lower() for ext in ['.pdf', '.xlsx', '.xls', '.doc', '.docx']):
                                if file_href.startswith('/'):
                                    file_url = base_url + file_href
                                elif file_href.startswith('http'):
                                    file_url = file_href
                                else:
                                    continue
                                
                                try:
                                    print(f"📄 Attempting to download file: {file_url}")
                                    file_result = await crawler.arun(url=file_url)
                                    if file_result.success and file_result.html:
                                        # Save the file content (this might be binary)
                                        file_name = os.path.basename(file_url).split('?')[0]
                                        if not file_name:
                                            file_name = f"file_{i}_{len(os.listdir(os.path.join(output_dir, 'files')))}"
                                        
                                        file_path = os.path.join(output_dir, "files", file_name)
                                        with open(file_path, "wb") as f:
                                            f.write(file_result.html.encode('utf-8'))
                                        print(f"✅ Downloaded: {file_name}")
                                except Exception as e:
                                    print(f"❌ Failed to download {file_url}: {e}")
                        
                        crawled_count += 1
                        
                        # Add delay between requests to be respectful
                        await asyncio.sleep(1)
                        
                    else:
                        print(f"❌ Failed to crawl: {link}")
                        
                except Exception as e:
                    print(f"❌ Error crawling {link}: {e}")
                    continue
            
            print(f"✅ Successfully crawled {crawled_count + 1} pages total")
            
        else:
            print("❌ Failed to crawl the main page")
            
    except Exception as e:
        print(f"❌ Crawling error: {e}")
        raise
    finally:
        # Properly close the crawler to avoid Windows asyncio warnings
        await crawler.close()

    print(f"✅ Crawl finished. All data saved in '{output_dir}/' directory.")
    return output_dir

def parse_html_metadata(output_dir):
    """Parses HTML files from the 'html' subdirectory and extracts metadata."""
    html_dir = os.path.join(output_dir, "html")
    if not os.path.exists(html_dir) or not os.listdir(html_dir):
        print(f"⚠️  No HTML files found to parse in '{html_dir}'.")
        return

    print(f"📄 Parsing HTML files from '{html_dir}'...")
    metadata_output = []
    
    for file in os.listdir(html_dir):
        if not file.endswith(".html"):
            continue

        path = os.path.join(html_dir, file)
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                soup = BeautifulSoup(f, "html.parser")
                title = soup.title.string.strip() if soup.title and soup.title.string else ""
                meta_tags = [meta.get("content", "") for meta in soup.find_all("meta") if meta.get("content")]
                images = [{"src": img.get("src", ""), "alt": img.get("alt", "")} for img in soup.find_all("img")]
                
                # Extract text content
                text_content = soup.get_text()
                clean_text = ' '.join(text_content.split())[:1000]  # First 1000 chars
                
                metadata_output.append({
                    "file": file, 
                    "title": title, 
                    "meta": meta_tags, 
                    "images": images,
                    "text_preview": clean_text
                })
                
        except Exception as e:
            print(f"❌ Failed to parse HTML {file}: {e}")

    if metadata_output:
        output_path = os.path.join(output_dir, "parsed_html_metadata.json")
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(metadata_output, f, indent=2, ensure_ascii=False)
        print(f"✅ Saved HTML metadata to '{output_path}'")

def parse_documents(output_dir):
    """Parses PDF and Excel files from the 'files' subdirectory."""
    file_dir = os.path.join(output_dir, "files")
    if not os.path.exists(file_dir) or not os.listdir(file_dir):
        print(f"⚠️  No downloaded documents found to parse in '{file_dir}'.")
        return

    print(f"📂 Parsing documents from '{file_dir}'...")
    docs_output = []
    
    for fname in os.listdir(file_dir):
        path = os.path.join(file_dir, fname)
        try:
            if fname.lower().endswith(".pdf"):
                try:
                    text = ""
                    with fitz.open(path) as doc:
                        for page in doc:
                            text += page.get_text()
                    docs_output.append({"file": fname, "type": "pdf", "content": text[:5000]})
                except Exception as e:
                    print(f"❌ Failed to parse PDF {fname}: {e}")
                    
            elif fname.lower().endswith((".xlsx", ".xls")):
                try:
                    wb = openpyxl.load_workbook(path, data_only=True)
                    rows = []
                    for sheet in wb.sheetnames:
                        ws = wb[sheet]
                        for row in ws.iter_rows(values_only=True):
                            if any(cell is not None for cell in row):
                                rows.append([str(cell) if cell is not None else "" for cell in row])
                    docs_output.append({"file": fname, "type": "xlsx", "content": rows[:50]})
                except Exception as e:
                    print(f"❌ Failed to parse Excel {fname}: {e}")
                    
        except Exception as e:
            print(f"❌ Failed to process document {fname}: {e}")

    if docs_output:
        output_path = os.path.join(output_dir, "parsed_documents.json")
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(docs_output, f, indent=2, ensure_ascii=False)
        print(f"✅ Saved parsed documents to '{output_path}'")

def main():
    """Main function with proper asyncio event loop handling for Windows."""
    install_playwright_browsers()
    
    # Set event loop policy for Windows to avoid warnings
    if sys.platform == "win32":
        asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
    
    try:
        # Create a new event loop
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        
        try:
            output_dir = loop.run_until_complete(run_the_crawl())
            parse_html_metadata(output_dir)
            parse_documents(output_dir)
            print("\n✨ All processing complete! ✨")
        finally:
            # Properly close the event loop
            loop.close()
            
    except Exception as e:
        print(f"🔥 A critical error occurred: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()